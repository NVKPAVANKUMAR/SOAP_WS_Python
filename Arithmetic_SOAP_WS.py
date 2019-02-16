import unittest
import HtmlTestRunner
import configparser
from openpyxl import load_workbook
from zeep import Client


def do_arithmetic(client, operation, num1, num2):
    if operation == "Add":
        return client.service.Add(num1, num2)
    elif operation == "Subtract":
        return client.service.Subtract(num1, num2)
    elif operation == "Multiply":
        return client.service.Multiply(num1, num2)
    elif operation == "Divide":
        return client.service.Divide(num1, num2)


def assert_result(actual_result, expected_result):
    assert actual_result == expected_result


def parse_excel_data(file_path, initial_cell, last_cell):
    book = load_workbook(file_path)
    sheet = book.active
    cells = sheet[initial_cell: last_cell]
    for c1, c2, c3, c4 in cells:
        return c1.value, c2.value, c3.value, c4.value


def parse_excel(file_path):
    rows = []
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
        rows.append(row)
    return rows


def write_to_excel(file_path, cell_position, message):
    book = load_workbook(file_path)
    sheet = book.active
    sheet.cell(row=cell_position, column=5).value = message
    book.save(file_path)


def config_parser(header, parameter):
    config = configparser.ConfigParser()
    config.read('configuration/config.ini')
    return config.get(header, parameter)


class Calculator_WS(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        global client
        client = ''
        cls.base_uri = config_parser("ConfigData", "calc_ws_url")
        client = Client(cls.base_uri)

    @staticmethod
    def test_01addition():
        actual_result = do_arithmetic(client, "Add", 5, 5)
        assert_result(actual_result, 10)

    @staticmethod
    def test_02subtraction():
        actual_result = do_arithmetic(client, "Subtract", 20, 10)
        assert_result(actual_result, 10)

    @staticmethod
    def test_03multiplication():
        actual_result = do_arithmetic(client, "Multiply", 20, 80)
        assert_result(actual_result, 1600)

    @staticmethod
    def test_04divide():
        actual_result = do_arithmetic(client, "Divide", 200, 10)
        assert_result(actual_result, 20)

    def test_parse_excel(self):
        file_path = config_parser("ConfigData", "file_path")
        out = parse_excel(file_path)
        for i in range(0, len(out)):
            actual_result = do_arithmetic(client, out[i][2].value, out[i][0].value, out[i][1].value)
            print actual_result
            if actual_result == out[i][3].value:
                write_to_excel(file_path, i + 2, "PASSED")
            else:
                write_to_excel(file_path, i + 2, "FAILED")


if __name__ == '__main__':
    unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output="example_dir"))
