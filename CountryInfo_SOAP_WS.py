import unittest

import HtmlTestRunner
import configparser
from openpyxl import load_workbook
from zeep import Client


def get_capital(client, countryISOcode):
    result = client.service.CapitalCity(countryISOcode)
    return result


def parse_excel(file_path):
    rows = []
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
        rows.append(row)
    return rows


def write_to_excel(file_path, cell_position, message):
    book = load_workbook(file_path)
    sheet = book.active
    sheet.cell(row=cell_position, column=4).value = message
    book.save(file_path)


def config_parser(header, parameter):
    config = configparser.ConfigParser()
    config.read('configuration/config.ini')
    return config.get(header, parameter)


class Country_WS_Tests(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        global client
        client = ''
        cls.base_uri = config_parser('ConfigData', 'country_ws_uri')
        client = Client(cls.base_uri)

    def test_ListOfContinentsByNameSoapRequest(self):
        result = client.service.ListOfContinentsByName()
        print result

    def test_get_capitalcity_ws(self):
        file_path = config_parser('ConfigData', 'countryData_file_path')
        out = parse_excel(file_path)
        for i in range(0, len(out)):
            actual_result = get_capital(client, out[i][1].value)
            if actual_result == out[i][2].value:
                write_to_excel(file_path, i + 2, "PASSED")
            else:
                write_to_excel(file_path, i + 2, "FAILED")


if __name__ == '__main__':
    unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output="example_dir"))
